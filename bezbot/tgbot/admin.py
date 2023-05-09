import csv
from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from django.http import HttpResponse
from django.core.exceptions import PermissionDenied
from django.contrib import admin

from django.contrib.auth.admin import UserAdmin
from .models import Users, Evacuation, Students, Missing


admin.site.site_title = 'Лицей 161'
admin.site.site_header = 'Лицей 161'

@admin.register(Users)
class UsersAdmin(UserAdmin):
    model = Users
    filter_horizontal = (
        "groups",
        "user_permissions",
    )
    list_display = ("username", "last_name", "first_name", "surename", "is_staff", "admin_verification")
    list_editable = ['admin_verification']
    fieldsets = (
        (None, {"fields": ("username", "password")}),
        (("Personal info"), {"fields": ("last_name", "first_name", "surename",)}),
        (
            ("Контакты"),
            {
                "fields": (
                    "chat_id",
                    "phone_number",
                    "email"
                ),
            }),(
            ("Permissions"),
            {
                "fields": (
                    "is_active",
                    "is_staff",
                    "is_superuser",
                    "admin_verification",
                    "groups",
                    "user_permissions",
                ),
            },
        ),
        (("Important dates"), {"fields": ("last_login", "date_joined")}),
    )
    add_fieldsets = (
        (
            None,
            {
                "classes": ("wide",),
                "fields": ("username", "password1", "password2"),
            },
        ),
    )


@admin.register(Students)
class StudentsAdmin(admin.ModelAdmin):
    model = Students
    search_fields = (
        'first_name',
        'last_name',
        'sure_name',
        'klass__name',
        'presence',

    )
    list_display = ("last_name", "first_name", 'sure_name', 'klass', 'presence',)
    list_editable = ('presence',)
    list_filter = ['klass']
    order = ('klass')


@admin.register(Evacuation)
class EvacuateAdmin(admin.ModelAdmin):
    model = Evacuation

    list_display = ('klass', "teacher", "status", 'before', 'amount', 'missing', 'klass_room', 'note', 'date')



    @admin.action(description="Экспортировать выбраное как XLSX")
    def download_missing_xlsx(modeladmin, request, queryset):
        if not request.user.is_staff:
            raise PermissionDenied
        opts = queryset.model._meta
        model = queryset.model

        wb = Workbook()
        ws = wb.active

        field_names = ['id', 'ФИО', 'Статус', 'Должно быть', 'В безопасности', 'Расхождение', 'Кабинет', 'Доп. информация', 'Дата']
        ws.append(field_names)

        for i, obj in enumerate(queryset):
            ws.append([i + 1, str(obj.teacher), obj.get_status_display(), str(obj.before), str(obj.amount), str(obj.missing), obj.klass_room, obj.note,
                             str(obj.date)])

        date = ws['F2'].value[:10]
        with NamedTemporaryFile(delete=False, mode='w+b') as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        response = HttpResponse(content=stream, content_type='application/ms-excel', )
        response['Content-Disposition'] = f'attachment; filename=evacuation-{date}.xlsx'

        return response

    actions = [ download_missing_xlsx]


@admin.register(Missing)
class MissingAdmin(admin.ModelAdmin):
    model = Missing

    list_display = ('student', "date", "reason",)
    

    @admin.action(description="Экспортировать выбраное как CSV")
    def download_missing_csv(modeladmin, request, queryset):
        if not request.user.is_staff:
            raise PermissionDenied
        opts = queryset.model._meta
        model = queryset.model
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment;filename=propuski.csv'
        writer = csv.writer(response)
        # field_names = [field.name for field in opts.fields]
        field_names = ['id', 'Фамилия', 'Имя', 'Класс', 'Причина', 'Дата']
        writer.writerow(field_names)
        # Write data rows
        for i, obj in enumerate(queryset):
            writer.writerow([i+1, obj.student.last_name, obj.student.first_name, obj.student.klass, obj.reason, str(obj.date)[:10]])
        return response

    @admin.action(description="Экспортировать выбраное как XLSX")
    def download_missing_xlsx(modeladmin, request, queryset):
        if not request.user.is_staff:
            raise PermissionDenied
        opts = queryset.model._meta
        model = queryset.model

        wb = Workbook()
        ws = wb.active

        field_names = ['id', 'Фамилия', 'Имя', 'Класс', 'Причина', 'Дата']
        ws.append(field_names)

        for i, obj in enumerate(queryset):
            ws.append([i+1, obj.student.last_name, obj.student.first_name, obj.student.klass.name, obj.reason, str(obj.date)[:10]])
        
        date = ws['F2'].value
        with NamedTemporaryFile(delete=False, mode='w+b') as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        response = HttpResponse(content=stream, content_type='application/ms-excel', )
        response['Content-Disposition'] = f'attachment; filename=propuski-{date}.xlsx'

        return response

    actions = [download_missing_csv, download_missing_xlsx]
